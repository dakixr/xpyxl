"""Process-isolated benchmark for xpyxl workflows that interact with OpenPyXL.

The default matrix covers generated saves, ``to_openpyxl()``, imported-only
workbooks, and mixed generated/imported workbooks through both the pure
OpenPyXL and hybrid engines::

    uv run python scripts/benchmark_openpyxl.py

Use ``--rows 1000 --import-rows 1000`` for a quick iteration-sized run.
"""

from __future__ import annotations

import argparse
import json
import resource
import statistics
import subprocess
import sys
import tempfile
import time
import zipfile
from collections.abc import Iterator, Mapping, Sequence
from dataclasses import asdict, dataclass
from datetime import date, timedelta
from pathlib import Path
from typing import Literal, cast, overload

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

import xpyxl as x
from xpyxl.engines import EngineName

__all__ = ["main"]

Scenario = Literal[
    "generated-save",
    "to-openpyxl",
    "import-openpyxl",
    "import-hybrid",
    "sparse-openpyxl",
    "sparse-hybrid",
    "mixed-openpyxl",
    "mixed-hybrid",
]

_SCENARIOS: tuple[Scenario, ...] = (
    "generated-save",
    "to-openpyxl",
    "import-openpyxl",
    "import-hybrid",
    "sparse-openpyxl",
    "sparse-hybrid",
    "mixed-openpyxl",
    "mixed-hybrid",
)
_COLUMNS = (
    "ID",
    "Booked",
    "Region",
    "Account",
    "Units",
    "Price",
    "Revenue",
    "Approved",
    "Forecast",
    "Notes",
)
_MIB = 1024 * 1024


class GeneratedRecords(Sequence[Mapping[str, object]]):
    """Repeatable deterministic records that do not retain cell objects."""

    def __init__(self, count: int, sheet: int) -> None:
        self._count = count
        self._sheet = sheet

    def __len__(self) -> int:
        return self._count

    @overload
    def __getitem__(self, index: int) -> Mapping[str, object]: ...

    @overload
    def __getitem__(self, index: slice) -> Sequence[Mapping[str, object]]: ...

    def __getitem__(
        self, index: int | slice
    ) -> Mapping[str, object] | Sequence[Mapping[str, object]]:
        if isinstance(index, slice):
            return tuple(self[item] for item in range(*index.indices(self._count)))
        if index < 0:
            index += self._count
        if index < 0 or index >= self._count:
            raise IndexError(index)

        sequence = self._sheet * self._count + index + 1
        units = 1 + sequence % 500
        price = 9.5 + sequence % 100
        excel_row = index + 7
        return {
            "ID": sequence,
            "Booked": date(2024, 1, 1) + timedelta(days=sequence % 730),
            "Region": ("AMER", "EMEA", "APAC")[sequence % 3],
            "Account": f"Account {sequence:06d}",
            "Units": units,
            "Price": price,
            "Revenue": units * price,
            "Approved": sequence % 7 != 0,
            "Forecast": f"=E{excel_row}*F{excel_row}",
            "Notes": f"Deterministic note {sequence}",
        }

    def __iter__(self) -> Iterator[Mapping[str, object]]:
        for index in range(self._count):
            yield self[index]


@dataclass(frozen=True)
class Result:
    scenario: str
    engine: str
    generated_rows: int
    imported_rows: int
    data_cells: int
    build_seconds: float
    operation_seconds: float
    total_seconds: float
    peak_rss_mib: float
    output_mib: float


def _peak_rss_mib() -> float:
    rss = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
    return rss / _MIB if sys.platform == "darwin" else rss / 1024


def _generated_sheet(rows: int, sheet_number: int) -> x.SheetNode:
    records = GeneratedRecords(rows, sheet_number)
    report_table = x.table(
        column_order=_COLUMNS,
        header_style=[x.bold, x.text_center, x.bg_primary],
        style=[x.table_bordered, x.table_banded, x.table_compact],
    )[records]
    return x.sheet(
        f"Generated {sheet_number + 1}",
        background_color="#F8FAFC",
        show_gridlines=False,
    )[
        x.row()[
            x.cell(colspan=len(_COLUMNS), style=[x.text_2xl, x.bold, x.bg_muted])[
                f"OpenPyXL interaction benchmark · {sheet_number + 1}"
            ]
        ],
        x.hstack(
            x.row(style=[x.bold, x.bg_info])["Rows", rows],
            x.row(style=[x.bold, x.bg_success])["Status", "Valid"],
            gap=2,
        ),
        x.space(height=8),
        report_table,
    ]


def _build_workbook(scenario: Scenario, rows: int, source: Path) -> x.Workbook:
    items: list[x.SheetNode | x.ImportedSheetNode] = []
    if scenario in ("generated-save", "to-openpyxl"):
        items.extend(_generated_sheet(rows, sheet) for sheet in range(2))
    elif scenario.startswith("mixed-"):
        items.append(_generated_sheet(rows, 0))

    if scenario.startswith("import-") or scenario.startswith("mixed-"):
        items.extend(
            [
                x.import_sheet(source, "Source 1", name="Imported 1"),
                x.import_sheet(source, "Source 2", name="Imported 2"),
            ]
        )
    elif scenario.startswith("sparse-"):
        items.append(x.import_sheet(source, "Sparse"))
    return x.workbook()[items]


def _engine_for(scenario: Scenario) -> EngineName:
    return "hybrid" if scenario.endswith("hybrid") else "openpyxl"


def _expected_sheets(scenario: Scenario) -> int:
    if scenario.startswith("mixed-"):
        return 3
    if scenario.startswith("sparse-"):
        return 1
    return 2


def _validate_xlsx(path: Path, expected_sheets: int) -> None:
    with zipfile.ZipFile(path) as archive:
        if archive.testzip() is not None:
            raise RuntimeError("Benchmark produced a corrupt XLSX archive")
        worksheets = [
            name
            for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        ]
    if len(worksheets) != expected_sheets:
        raise RuntimeError(f"Expected {expected_sheets} sheets, got {len(worksheets)}")


def _worker(scenario: Scenario, rows: int, import_rows: int, source: Path) -> Result:
    started = time.perf_counter()
    workbook = _build_workbook(scenario, rows, source)
    built = time.perf_counter()
    output_mib = 0.0

    if scenario == "to-openpyxl":
        rendered = workbook.to_openpyxl()
        if rendered.sheetnames != ["Generated 1", "Generated 2"]:
            raise RuntimeError("to_openpyxl produced unexpected sheets")
        if rendered["Generated 2"].cell(rows + 4, 1).value != rows * 2:
            raise RuntimeError("to_openpyxl produced unexpected data")
        completed = time.perf_counter()
    else:
        with tempfile.TemporaryDirectory(prefix="xpyxl-openpyxl-benchmark-") as tmp:
            output = Path(tmp) / "result.xlsx"
            workbook.save(output, engine=_engine_for(scenario))
            completed = time.perf_counter()
            _validate_xlsx(output, _expected_sheets(scenario))
            output_mib = output.stat().st_size / _MIB

    generated_sheets = 2 if scenario in ("generated-save", "to-openpyxl") else int(
        scenario.startswith("mixed-")
    )
    imported_sheets = 2 if scenario.startswith(("import-", "mixed-")) else 0
    generated_row_count = rows * generated_sheets
    imported_row_count = import_rows * imported_sheets
    data_cells = (generated_row_count + imported_row_count) * len(_COLUMNS)
    if scenario.startswith("sparse-"):
        imported_row_count = import_rows * 4
        data_cells = 2
    return Result(
        scenario=scenario,
        engine=_engine_for(scenario),
        generated_rows=generated_row_count,
        imported_rows=imported_row_count,
        data_cells=data_cells,
        build_seconds=built - started,
        operation_seconds=completed - built,
        total_seconds=completed - started,
        peak_rss_mib=_peak_rss_mib(),
        output_mib=output_mib,
    )


def _create_import_source(path: Path, rows: int) -> None:
    workbook = openpyxl.Workbook()
    side = Side(style="thin", color="FFD1D5DB")
    border = Border(left=side, right=side, top=side, bottom=side)
    fills = (PatternFill("solid", fgColor="FFF8FAFC"), PatternFill())

    for sheet_number in range(2):
        sheet = workbook.active if sheet_number == 0 else workbook.create_sheet()
        if sheet is None:
            raise RuntimeError("Expected an active source sheet")
        sheet.title = f"Source {sheet_number + 1}"
        sheet.merge_cells("A1:J1")
        sheet["A1"] = f"Import source {sheet_number + 1}"
        sheet["A1"].font = Font(size=18, bold=True, color="FF123456")
        sheet["A1"].fill = PatternFill("solid", fgColor="FFE2E8F0")
        for column, value in enumerate(_COLUMNS, 1):
            cell = sheet.cell(3, column, value)
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor="FF2563EB")

        records = GeneratedRecords(rows, sheet_number)
        for index in range(rows):
            row = index + 4
            values = tuple(records[index].values())
            for column, value in enumerate(values, 1):
                cell = sheet.cell(
                    row,
                    column,
                    cast(str | int | float | bool | date, value),
                )
                cell.border = border
                cell.fill = fills[index % 2]
                if column in (5, 6, 7):
                    cell.number_format = "#,##0.00"
                if column == 10:
                    cell.alignment = Alignment(wrap_text=True)

        sheet.freeze_panes = "A4"
        sheet.auto_filter.ref = f"A3:J{rows + 3}"
        sheet.column_dimensions["D"].width = 22
        sheet.column_dimensions["J"].width = 30
        table = Table(
            displayName=f"SourceTable{sheet_number + 1}", ref=f"A3:J{rows + 3}"
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True
        )
        sheet.add_table(table)

    sparse = workbook.create_sheet("Sparse")
    sparse["A1"] = "Start"
    sparse.cell(row=rows * 4, column=20, value="End")
    workbook.save(path)


def _run_child(
    scenario: Scenario,
    args: argparse.Namespace,
    source: Path,
) -> Result:
    command = [
        sys.executable,
        str(Path(__file__).resolve()),
        "--worker",
        "--scenario",
        scenario,
        "--rows",
        str(args.rows),
        "--import-rows",
        str(args.import_rows),
        "--source",
        str(source),
    ]
    completed = subprocess.run(command, check=True, capture_output=True, text=True)
    return Result(**json.loads(completed.stdout.strip().splitlines()[-1]))


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--scenario", choices=("all", *_SCENARIOS), default="all")
    parser.add_argument("--rows", type=int, default=5_000)
    parser.add_argument("--import-rows", type=int, default=5_000)
    parser.add_argument("--runs", type=int, default=1)
    parser.add_argument("--json", action="store_true")
    parser.add_argument("--worker", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--source", type=Path, help=argparse.SUPPRESS)
    args = parser.parse_args()
    if args.rows < 1 or args.import_rows < 1 or args.runs < 1:
        parser.error("rows, import-rows, and runs must be positive")
    if args.worker and args.scenario == "all":
        parser.error("worker requires one concrete scenario")
    if args.worker and args.source is None:
        parser.error("worker requires --source")
    return args


def main() -> None:
    args = _parse_args()
    if args.worker:
        scenario = cast(Scenario, args.scenario)
        print(json.dumps(asdict(_worker(scenario, args.rows, args.import_rows, args.source))))
        return

    scenarios = _SCENARIOS if args.scenario == "all" else (cast(Scenario, args.scenario),)
    with tempfile.TemporaryDirectory(prefix="xpyxl-openpyxl-source-") as tmp:
        source = Path(tmp) / "source.xlsx"
        if any(
            scenario not in ("generated-save", "to-openpyxl")
            for scenario in scenarios
        ):
            _create_import_source(source, args.import_rows)
        results: list[dict[str, object]] = []
        for scenario in scenarios:
            runs = [_run_child(scenario, args, source) for _ in range(args.runs)]
            representative = min(runs, key=lambda item: item.total_seconds)
            results.append(
                {
                    **asdict(representative),
                    "runs": args.runs,
                    "median_total_seconds": statistics.median(
                        item.total_seconds for item in runs
                    ),
                    "median_peak_rss_mib": statistics.median(
                        item.peak_rss_mib for item in runs
                    ),
                }
            )

    if args.json:
        print(json.dumps(results, indent=2, sort_keys=True))
        return
    for result in results:
        print(
            f"{result['scenario']:<17} {result['engine']:<9} "
            f"{result['median_total_seconds']:>8.2f}s  "
            f"{result['median_peak_rss_mib']:>8.1f} MiB  "
            f"{result['data_cells']:>9,} cells"
        )


if __name__ == "__main__":
    main()
