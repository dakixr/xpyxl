from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

import xpyxl as x


def test_cell_builder_rejects_invalid_colspan() -> None:
    with pytest.raises(ValueError, match="Cell colspan must be >= 1"):
        x.cell(colspan=0)["bad"]


def test_cell_builder_rejects_invalid_rowspan() -> None:
    with pytest.raises(ValueError, match="Cell rowspan must be >= 1"):
        x.cell(rowspan=0)["bad"]


def test_raw_scalar_row_cells_remain_unmerged() -> None:
    workbook = x.workbook()[x.sheet("Plain")[x.row()["A", "B", "C"]]]

    result = workbook.to_openpyxl()
    ws = result["Plain"]

    assert list(ws.merged_cells.ranges) == []
    assert ws["A1"].value == "A"
    assert ws["B1"].value == "B"
    assert ws["C1"].value == "C"


@pytest.mark.parametrize("engine", ["openpyxl", "xlsxwriter", "hybrid"])
def test_generated_merged_cells_round_trip(engine: str) -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / f"{engine}.xlsx"

        workbook = x.workbook()[
            x.sheet("Merged")[
                x.row()[
                    x.cell(style=[x.bold, x.bg_warning, x.text_center], colspan=3)[
                        "Merged Title"
                    ]
                ],
                x.row()[
                    x.cell(style=[x.bg_info], rowspan=2)["Region"],
                    "Q1",
                    "Q2",
                ],
                x.row()["North", 10, 20],
            ]
        ]

        workbook.save(output_path, engine=engine)  # type: ignore[arg-type]

        result = openpyxl.load_workbook(output_path)
        ws = result["Merged"]
        merged_ranges = {str(merged_range) for merged_range in ws.merged_cells.ranges}

        assert merged_ranges == {"A1:C1", "A2:A3"}
        assert ws["A1"].value == "Merged Title"
        assert ws["A1"].font.bold
        assert ws["A2"].value == "Region"
        assert ws["B2"].value == "Q1"
        assert ws["C2"].value == "Q2"
        assert ws["B3"].value == "North"
        assert ws["C3"].value == 10
        assert ws["D3"].value == 20


def test_rowspan_conflict_with_block_content_raises() -> None:
    workbook = x.workbook()[
        x.sheet("Conflict")[
            x.row()[x.cell(rowspan=2)["Merged"]],
            x.col()["Overlaps"],
        ]
    ]

    with pytest.raises(ValueError, match="Merged cells cannot overlap existing content"):
        workbook.save(engine="openpyxl")


def test_merged_cells_inside_tables_raise() -> None:
    workbook = x.workbook()[
        x.sheet("Table")[
            x.table()[[x.row()[x.cell(colspan=2)["Not allowed"]]]],
        ]
    ]

    with pytest.raises(ValueError, match="Merged cells are not supported inside tables"):
        workbook.save(engine="openpyxl")


def test_generated_html_renders_colspan_and_rowspan() -> None:
    workbook = x.workbook()[
        x.sheet("Merged")[
            x.row()[x.cell(colspan=3)["Merged Title"]],
            x.row()[x.cell(rowspan=2)["Region"], "Q1"],
            x.row()["North"],
        ]
    ]

    result = workbook.save(engine="html")
    assert isinstance(result, bytes)
    html = result.decode("utf-8")

    assert 'colspan="3"' in html
    assert 'rowspan="2"' in html
    assert html.count("Merged Title") == 1
    assert html.count("Region") == 1


def test_wrapping_merged_row_in_vstack_preserves_layout() -> None:
    direct = x.workbook()[
        x.sheet("Direct")[
            x.row()[x.cell(rowspan=2)["Region"], "Q1"],
            x.row()["North"],
        ]
    ].to_openpyxl()["Direct"]

    wrapped = x.workbook()[
        x.sheet("Wrapped")[
            x.vstack(x.row()[x.cell(rowspan=2)["Region"], "Q1"]),
            x.row()["North"],
        ]
    ].to_openpyxl()["Wrapped"]

    for row in range(1, 4):
        for col in range(1, 4):
            assert wrapped.cell(row=row, column=col).value == direct.cell(
                row=row, column=col
            ).value


def test_explicit_row_height_applies_to_each_row_in_rowspan() -> None:
    workbook = x.workbook()[
        x.sheet("Heights")[
            x.row()[x.cell(style=[x.row_height(40)], rowspan=2)["Merged"]],
            x.row()["Next"],
        ]
    ]

    ws = workbook.to_openpyxl()["Heights"]

    assert ws.row_dimensions[1].height == 40
    assert ws.row_dimensions[2].height == 40


def test_imported_html_renders_existing_merged_ranges() -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        source_path = Path(tmpdir) / "source.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            raise RuntimeError("Expected an active worksheet")
        ws.title = "Source"
        ws["A1"] = "Merged Title"
        ws.merge_cells("A1:C1")
        ws["A2"] = "Region"
        ws.merge_cells("A2:A3")
        ws["B2"] = "Q1"
        ws["B3"] = "North"
        wb.save(source_path)

        workbook = x.workbook()[x.import_sheet(source_path, "Source", name="Imported")]
        result = workbook.save(engine="html")

        assert isinstance(result, bytes)
        html = result.decode("utf-8")
        assert 'colspan="3"' in html
        assert 'rowspan="2"' in html
        assert "Merged Title" in html
        assert "Region" in html
