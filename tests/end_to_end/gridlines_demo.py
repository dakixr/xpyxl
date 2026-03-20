"""Manual visual demo for sheet-level gridline visibility."""

from __future__ import annotations

from pathlib import Path

import openpyxl

import xpyxl as x


def _build_template(path: Path) -> None:
    path.parent.mkdir(exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Expected an active worksheet in template workbook")

    ws.title = "HiddenGridTemplate"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Imported template with hidden gridlines"
    ws["A3"] = "Row"
    ws["B3"] = "Value"
    ws["A4"] = "North"
    ws["B4"] = 120
    ws["A5"] = "South"
    ws["B5"] = 95
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    wb.save(path)


def _generated_sheet() -> x.SheetNode:
    return x.sheet("Generated Gridlines", show_gridlines=False)[
        x.row(style=[x.text_2xl, x.bold])["Generated sheet"],
        x.row(style=[x.text_sm, x.text_gray])[
            "Gridlines should be hidden here even though no borders are applied."
        ],
        x.space(),
        x.row()["Metric", "Q1", "Q2", "Q3"],
        x.row()["Revenue", 120, 128, 135],
        x.row()["Margin", "32%", "35%", "37%"],
        x.space(),
        x.row(style=[x.text_sm, x.text_gray])[
            "Open the xlsx/html outputs in .testing and confirm the sheet looks clean by default."
        ],
    ]


def _imported_sheet(template_path: Path) -> x.ImportedSheetNode:
    return x.import_sheet(
        template_path,
        "HiddenGridTemplate",
        name="Imported Gridlines",
    )


def build_workbook() -> list[x.SheetNode | x.ImportedSheetNode]:
    project_root = Path(__file__).resolve().parent.parent
    template_path = project_root / ".testing" / "gridlines-template.xlsx"
    _build_template(template_path)
    return [_generated_sheet(), _imported_sheet(template_path)]
