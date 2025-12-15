from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

import xpyxl as x


def _build_template(path: Path) -> None:
    """Create a small template workbook with styles and merges."""

    path.parent.mkdir(exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Expected an active worksheet in template workbook")
    ws.title = "Template"

    ws["A1"] = "Template Cover"
    ws["A1"].style = "Title"
    ws["A1"].fill = PatternFill(
        start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
    )
    ws.merge_cells("A1:C1")

    ws["F1"].fill = PatternFill(
        start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
    )
    ws.merge_cells("F1:F12")

    ws["A3"] = "Notes"
    ws["A4"] = "Static content from Excel"

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.row_dimensions[1].height = 28

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = "A3:C10"

    wb.save(path)


def build_workbook():
    project_root = Path(__file__).resolve().parent.parent
    template_path = project_root / ".testing" / "import-template.xlsx"
    _build_template(template_path)
    return [x.import_sheet(template_path, "Template", name="Imported Template")]
