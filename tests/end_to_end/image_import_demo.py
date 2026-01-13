"""Demo for importing sheets with images and charts.

This demo creates a template with embedded images and charts, then imports it
using xpyxl's import_sheet functionality. You can manually verify that images
and charts are preserved in the output files.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill

import xpyxl as x

# Check if Pillow is available for image creation
PIL_AVAILABLE = True
PILImage: Any = None
try:
    from PIL import Image as _PILImage

    PILImage = _PILImage
except ImportError:
    PIL_AVAILABLE = False


def _create_colored_image(color: str, size: tuple[int, int] = (120, 80)) -> BytesIO:
    """Create a simple colored PNG image."""
    if not PIL_AVAILABLE or PILImage is None:
        raise ImportError("Pillow is required to create images")

    pil_img = PILImage.new("RGB", size, color=color)
    img_bytes = BytesIO()
    pil_img.save(img_bytes, format="PNG")
    img_bytes.seek(0)
    return img_bytes


def _build_template_with_images(path: Path) -> None:
    """Create a template workbook containing images and a chart."""
    path.parent.mkdir(exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Expected an active worksheet")
    ws.title = "ImageDemo"

    # Title
    ws["A1"] = "Image & Chart Import Demo"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    ws.merge_cells("A1:F1")

    # Add description
    ws["A3"] = "This sheet contains embedded images and a chart."
    ws["A4"] = "When imported, all visual elements should be preserved."

    # Add colored images at different positions
    colors = [
        ("red", "B6"),
        ("green", "D6"),
        ("blue", "F6"),
    ]

    for color, anchor in colors:
        img_bytes = _create_colored_image(color)
        img = Image(img_bytes)
        img.anchor = anchor
        ws.add_image(img)

    # Add labels for images
    ws["B5"] = "Red Image"
    ws["D5"] = "Green Image"
    ws["F5"] = "Blue Image"

    # Add data for chart
    ws["A12"] = "Chart Data"
    ws["A12"].font = Font(bold=True)
    data_items = [
        ("Q1", 120),
        ("Q2", 180),
        ("Q3", 150),
        ("Q4", 220),
    ]
    for i, (label, value) in enumerate(data_items):
        ws.cell(row=13 + i, column=1, value=label)
        ws.cell(row=13 + i, column=2, value=value)

    # Create bar chart
    chart = BarChart()
    chart.title = "Quarterly Sales"
    chart.style = 10
    data = Reference(ws, min_col=2, min_row=13, max_row=16)
    categories = Reference(ws, min_col=1, min_row=13, max_row=16)
    chart.add_data(data)
    chart.set_categories(categories)
    ws.add_chart(chart, "D12")  # type: ignore[arg-type]

    # Set column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 5
    ws.column_dimensions["F"].width = 15

    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[6].height = 60

    wb.save(path)


def build_workbook() -> list[x.ImportedSheetNode]:
    """Build and return imported sheets with images for the combined demo."""
    if not PIL_AVAILABLE:
        # Return an empty list if Pillow is not available
        # This allows the demo to be skipped gracefully
        return []

    project_root = Path(__file__).resolve().parent.parent.parent
    template_path = project_root / ".testing" / "image-demo-template.xlsx"

    _build_template_with_images(template_path)

    return [x.import_sheet(template_path, "ImageDemo", name="Image Demo")]
