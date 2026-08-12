from __future__ import annotations

from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import cast

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.styles.numbers import is_date_format
from openpyxl.worksheet.worksheet import Worksheet

import xpyxl as x
from xpyxl.engines._openpyxl_compat import CompiledStyleCache, populated_cells


def test_openpyxl_compatibility_helpers_fail_open() -> None:
    cell = cast(Cell, object())
    cache: CompiledStyleCache[str] = CompiledStyleCache()

    class PublicOnlySheet:
        def iter_rows(self) -> tuple[tuple[Cell, ...], ...]:
            return ((cell,),)

    assert cache.apply(cell, "style") is False
    cache.capture(cell, "style")
    assert list(populated_cells(cast(Worksheet, PublicOnlySheet()))) == [cell]


def test_compiled_styles_remain_independently_mutable() -> None:
    workbook = x.workbook()[x.sheet("S")[x.row(style=[x.bold])["A", "B"]]]

    sheet = workbook.to_openpyxl()["S"]

    assert sheet["A1"]._style == sheet["B1"]._style
    assert sheet["A1"]._style is not sheet["B1"]._style
    sheet["A1"].font = Font(bold=False)
    assert sheet["A1"].font.bold is False
    assert sheet["B1"].font.bold is True


def test_compiled_styles_preserve_automatic_date_formats() -> None:
    workbook = x.workbook()[
        x.sheet("S")[
            x.row()[date(2026, 8, 12), date(2026, 8, 13)],
            x.row()[
                x.cell(style=[x.date_short])[date(2026, 8, 14)],
                x.cell(style=[x.date_short])[date(2026, 8, 15)],
            ],
            x.row()[x.cell(style=[x.currency_usd])[date(2026, 8, 16)]],
        ]
    ]

    sheet = workbook.to_openpyxl()["S"]

    assert is_date_format(sheet["A1"].number_format)
    assert sheet["A1"].number_format == sheet["B1"].number_format
    assert sheet["A2"].number_format == sheet["B2"].number_format
    assert not is_date_format(sheet["A3"].number_format)


def test_imported_compiled_styles_remain_independently_mutable() -> None:
    with TemporaryDirectory() as tmpdir:
        source_path = Path(tmpdir) / "source.xlsx"
        source = openpyxl.Workbook()
        source_sheet = source.active
        assert source_sheet is not None
        source_sheet.title = "Source"
        shared_font = Font(bold=True, color="FF123456")
        source_sheet["A1"] = "A"
        source_sheet["A1"].font = shared_font
        source_sheet["A2"] = "B"
        source_sheet["A2"].font = shared_font
        source.save(source_path)

        imported = x.workbook()[x.import_sheet(source_path, "Source")].to_openpyxl()

    sheet = imported["Source"]
    assert sheet["A1"]._style == sheet["A2"]._style
    assert sheet["A1"]._style is not sheet["A2"]._style
    sheet["A1"].font = Font(bold=False)
    assert sheet["A1"].font.bold is False
    assert sheet["A2"].font.bold is True


def test_sparse_import_does_not_materialize_empty_rectangle() -> None:
    with TemporaryDirectory() as tmpdir:
        source_path = Path(tmpdir) / "sparse.xlsx"
        source = openpyxl.Workbook()
        source_sheet = source.active
        assert source_sheet is not None
        source_sheet.title = "Sparse"
        source_sheet["A1"] = "Start"
        source_sheet["T2000"] = "End"
        source.save(source_path)

        imported = x.workbook()[x.import_sheet(source_path, "Sparse")].to_openpyxl()

    sheet = imported["Sparse"]
    assert sheet.max_row == 2000
    assert sheet.max_column == 20
    assert sheet["A1"].value == "Start"
    assert sheet["T2000"].value == "End"
    assert len(sheet._cells) == 2
