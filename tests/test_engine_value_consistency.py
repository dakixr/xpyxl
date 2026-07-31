"""Cross-engine consistency tests for cell value handling."""

from __future__ import annotations

import tempfile
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

import xpyxl as x

ENGINES = ["openpyxl", "xlsxwriter", "hybrid"]


@dataclass(frozen=True)
class Point:
    x: int
    y: int


@pytest.mark.parametrize("engine", ENGINES)
def test_decimal_values_are_numeric(engine: str) -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / f"{engine}.xlsx"
        workbook = x.workbook()[x.sheet("S")[x.row()[Decimal("1.5")]]]
        workbook.save(output_path, engine=engine)  # type: ignore[arg-type]

        ws = openpyxl.load_workbook(output_path)["S"]
        assert ws["A1"].value == 1.5
        assert not isinstance(ws["A1"].value, str)


@pytest.mark.parametrize("engine", ENGINES)
def test_formula_strings_become_formulas(engine: str) -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / f"{engine}.xlsx"
        workbook = x.workbook()[x.sheet("S")[x.row()["=1+1"]]]
        workbook.save(output_path, engine=engine)  # type: ignore[arg-type]

        ws = openpyxl.load_workbook(output_path)["S"]
        assert ws["A1"].data_type == "f"
        assert ws["A1"].value == "=1+1"


@pytest.mark.parametrize("engine", ENGINES)
def test_unsupported_values_degrade_to_string(engine: str) -> None:
    point = Point(1, 2)
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / f"{engine}.xlsx"
        workbook = x.workbook()[x.sheet("S")[x.row()[point]]]
        workbook.save(output_path, engine=engine)  # type: ignore[arg-type]

        ws = openpyxl.load_workbook(output_path)["S"]
        assert ws["A1"].value == str(point)


@pytest.mark.parametrize("engine", ENGINES)
def test_bytes_values_decode_to_text(engine: str) -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / f"{engine}.xlsx"
        workbook = x.workbook()[x.sheet("S")[x.row()[b"abc"]]]
        workbook.save(output_path, engine=engine)  # type: ignore[arg-type]

        ws = openpyxl.load_workbook(output_path)["S"]
        assert ws["A1"].value == "abc"


def test_html_bytes_values_decode_to_text() -> None:
    result = x.workbook()[x.sheet("S")[x.row()[b"abc"]]].save(engine="html")
    assert isinstance(result, bytes)
    html = result.decode("utf-8")
    assert "abc" in html
    assert "b&#x27;abc&#x27;" not in html


@pytest.mark.parametrize("engine", ENGINES)
def test_merged_cell_values_follow_same_rules(engine: str) -> None:
    point = Point(3, 4)
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / f"{engine}.xlsx"
        workbook = x.workbook()[
            x.sheet("S")[
                x.row()[x.cell(colspan=2)[point]],
                x.row()[x.cell(colspan=2)["=1+1"]],
                x.row()[x.cell(colspan=2)[Decimal("2.5")]],
            ]
        ]
        workbook.save(output_path, engine=engine)  # type: ignore[arg-type]

        ws = openpyxl.load_workbook(output_path)["S"]
        assert ws["A1"].value == str(point)
        assert ws["A2"].data_type == "f"
        assert ws["A3"].value == 2.5
