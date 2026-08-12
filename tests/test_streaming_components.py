from __future__ import annotations

import tempfile
from collections.abc import Iterator, Sequence
from pathlib import Path
from typing import overload

import openpyxl
import pytest

import xpyxl as x


class _Components(Sequence[x.Node]):
    def __init__(self, count: int) -> None:
        self.component_count = count
        self.generated = 0

    def __len__(self) -> int:
        return self.component_count

    @overload
    def __getitem__(self, index: int) -> x.Node: ...

    @overload
    def __getitem__(self, index: slice) -> Sequence[x.Node]: ...

    def __getitem__(self, index: int | slice) -> x.Node | Sequence[x.Node]:
        if isinstance(index, slice):
            return tuple(
                self[item] for item in range(*index.indices(self.component_count))
            )
        if index < 0:
            index += self.component_count
        if index < 0 or index >= self.component_count:
            raise IndexError(index)
        self.generated += 1
        return x.vstack(
            x.row()[x.cell(colspan=3, style=[x.bold, x.bg_muted])[f"Block {index}"]],
            x.hstack(
                x.col(style=[x.text_sm])["Label", "Value"],
                x.col(style=[x.number_comma])[index, index * 2],
                gap=1,
            ),
            x.space(height=24 + index),
            gap=1,
        )

    def __iter__(self) -> Iterator[x.Node]:
        for index in range(self.component_count):
            yield self[index]


def test_repeatable_sheet_component_sequence_remains_lazy() -> None:
    components = _Components(3)

    workbook = x.workbook()[x.sheet("Lazy")[components]]

    assert components.generated == 0
    sheet = workbook._node.sheets[0]
    assert isinstance(sheet, x.SheetNode)
    assert not isinstance(sheet.items, tuple)


@pytest.mark.parametrize("engine", ["xlsxwriter", "hybrid"])
def test_streamed_components_match_materialized_layout(engine: str) -> None:
    lazy_components = _Components(3)
    materialized_components = tuple(_Components(3))

    with tempfile.TemporaryDirectory() as tmpdir:
        lazy_path = Path(tmpdir) / "lazy.xlsx"
        materialized_path = Path(tmpdir) / "materialized.xlsx"
        x.workbook()[x.sheet("S")[lazy_components]].save(
            lazy_path, engine=engine  # type: ignore[arg-type]
        )
        x.workbook()[x.sheet("S")[materialized_components]].save(
            materialized_path, engine=engine  # type: ignore[arg-type]
        )
        lazy = openpyxl.load_workbook(lazy_path)["S"]
        materialized = openpyxl.load_workbook(materialized_path)["S"]

    assert lazy.max_row == materialized.max_row
    assert lazy.max_column == materialized.max_column
    assert lazy.merged_cells.ranges == materialized.merged_cells.ranges
    for row in range(1, lazy.max_row + 1):
        assert lazy.row_dimensions[row].height == materialized.row_dimensions[row].height
        for col in range(1, lazy.max_column + 1):
            lazy_cell = lazy.cell(row, col)
            materialized_cell = materialized.cell(row, col)
            assert lazy_cell.value == materialized_cell.value
            assert lazy_cell.style_id == materialized_cell.style_id


@pytest.mark.parametrize("engine", ["xlsxwriter", "hybrid"])
def test_streamed_spacer_only_row_keeps_its_height(engine: str) -> None:
    components = _Components(1)

    with tempfile.TemporaryDirectory() as tmpdir:
        output = Path(tmpdir) / "spacer.xlsx"
        x.workbook()[x.sheet("S")[components]].save(
            output, engine=engine  # type: ignore[arg-type]
        )
        sheet = openpyxl.load_workbook(output)["S"]

    spacer_rows = [
        row for row, dimension in sheet.row_dimensions.items() if dimension.height == 24
    ]
    assert spacer_rows == [6]


@pytest.mark.parametrize("engine", ["xlsxwriter", "hybrid"])
def test_streamed_components_keep_sheet_background(engine: str) -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        output = Path(tmpdir) / "background.xlsx"
        x.workbook()[
            x.sheet("S", background_color="#F8FAFC")[_Components(1)]
        ].save(output, engine=engine)  # type: ignore[arg-type]
        sheet = openpyxl.load_workbook(output)["S"]

    assert sheet["A3"].fill.fgColor.rgb == "FFF8FAFC"
    assert len(sheet.conditional_formatting) == 1


@pytest.mark.parametrize("engine", ["xlsxwriter", "hybrid"])
def test_lazy_rowspan_components_use_random_access_fallback(engine: str) -> None:
    class Rowspans(Sequence[x.Node]):
        def __len__(self) -> int:
            return 2

        @overload
        def __getitem__(self, index: int) -> x.Node: ...

        @overload
        def __getitem__(self, index: slice) -> Sequence[x.Node]: ...

        def __getitem__(self, index: int | slice) -> x.Node | Sequence[x.Node]:
            if isinstance(index, slice):
                return tuple(self[item] for item in range(*index.indices(len(self))))
            if index == 0:
                return x.row()[x.cell(rowspan=2)["Tall"], "Alongside"]
            if index == 1:
                return x.row()["Flows around"]
            raise IndexError(index)

    with tempfile.TemporaryDirectory() as tmpdir:
        output = Path(tmpdir) / "rowspan.xlsx"
        x.workbook()[x.sheet("S")[Rowspans()]].save(
            output, engine=engine  # type: ignore[arg-type]
        )
        sheet = openpyxl.load_workbook(output)["S"]

    assert {str(item) for item in sheet.merged_cells.ranges} == {"A1:A2"}
    assert sheet["A1"].value == "Tall"
    assert sheet["B1"].value == "Alongside"
    assert sheet["B2"].value == "Flows around"
