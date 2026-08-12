"""Tests for sheet-level background_color rendering."""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

import xpyxl as x

ENGINES = ["openpyxl", "xlsxwriter", "hybrid"]


@pytest.mark.parametrize("engine", ENGINES)
def test_background_color_applies_to_populated_cells(engine: str) -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / f"{engine}.xlsx"
        workbook = x.workbook()[
            x.sheet("S", background_color="#F8FAFC")[
                x.row()["plain"],
                x.row()[x.cell(style=[x.bg_warning])["own-fill"]],
            ]
        ]
        workbook.save(output_path, engine=engine)  # type: ignore[arg-type]

        ws = openpyxl.load_workbook(output_path)["S"]

        assert ws["A1"].fill.fgColor.rgb == "FFF8FAFC"
        # An explicit cell fill still wins over the sheet background.
        assert ws["A2"].fill.fgColor.rgb == "FFB45309"


def test_openpyxl_background_does_not_materialize_blank_range() -> None:
    workbook = x.workbook()[
        x.sheet("S", background_color="#F8FAFC")[x.row()["plain"]]
    ]

    sheet = workbook.to_openpyxl()["S"]

    assert sheet.max_row == 1
    assert sheet.max_column == 1
    assert sheet["A1"].fill.fgColor.rgb == "FFF8FAFC"
    assert len(sheet.conditional_formatting) == 1
