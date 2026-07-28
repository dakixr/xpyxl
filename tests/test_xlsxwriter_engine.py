from __future__ import annotations

import tempfile
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pytest

import xpyxl as x


@pytest.mark.parametrize("engine", ["openpyxl", "xlsxwriter", "hybrid"])
def test_formula_strings_are_written_as_formulas(engine: str) -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / f"{engine}.xlsx"
        workbook = x.workbook()[x.sheet("Formulas")[x.row()["=1+1"]]]

        workbook.save(output_path, engine=engine)  # type: ignore[arg-type]

        cell = openpyxl.load_workbook(output_path, data_only=False)["Formulas"]["A1"]
        assert cell.value == "=1+1"
        assert cell.data_type == "f"


@pytest.mark.parametrize("engine", ["openpyxl", "xlsxwriter", "hybrid"])
def test_bare_equals_sign_is_written_as_text(engine: str) -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / f"{engine}.xlsx"
        workbook = x.workbook()[x.sheet("Text")[x.row()["="]]]

        workbook.save(output_path, engine=engine)  # type: ignore[arg-type]

        cell = openpyxl.load_workbook(output_path, data_only=False)["Text"]["A1"]
        assert cell.value == "="
        assert cell.data_type == "s"


def test_xlsxwriter_applies_custom_date_formats() -> None:
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "output.xlsx"
        custom_date = x.Style(name="custom_date", number_format="dd/mm/yyyy")
        custom_datetime = x.Style(
            name="custom_datetime", number_format="dd/mm/yyyy hh:mm"
        )

        workbook = x.workbook()[
            x.sheet("Dates")[
                x.row()[
                    x.cell(style=[custom_date])[date(2025, 1, 2)],
                    x.cell(style=[custom_datetime])[datetime(2025, 1, 2, 14, 30)],
                ]
            ]
        ]

        workbook.save(output_path, engine="xlsxwriter")

        result_wb = openpyxl.load_workbook(output_path)
        ws = result_wb["Dates"]

        assert ws["A1"].number_format == "dd/mm/yyyy"
        assert ws["B1"].number_format == "dd/mm/yyyy hh:mm"
        assert ws["A1"].is_date
        assert ws["B1"].is_date
        assert not isinstance(ws["A1"].value, str)
        assert not isinstance(ws["B1"].value, str)
