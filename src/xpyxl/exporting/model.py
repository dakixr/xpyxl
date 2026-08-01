"""Renderer-neutral workbook layout used by document exporters."""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from ..engines.base import EffectiveStyle, Engine, SaveTarget, normalize_cell_value
from ..nodes import WorkbookNode
from ..render import render_sheet
from ..styles import BorderStyleName, normalize_hex

__all__ = [
    "CellLayout",
    "SheetLayout",
    "WorkbookLayout",
    "build_workbook_layout",
]

_DEFAULT_FONT_NAME = "Calibri"
_DEFAULT_FONT_SIZE = 11.0
_DEFAULT_TEXT_COLOR = "#000000"
_DEFAULT_BORDER_COLOR = "#000000"


@dataclass(frozen=True)
class CellLayout:
    """A cell positioned in a renderer-neutral sheet grid."""

    row: int
    col: int
    value: object
    style: EffectiveStyle
    border_fallback_color: str
    rowspan: int = 1
    colspan: int = 1


@dataclass
class SheetLayout:
    """Resolved cells and dimensions for one sheet."""

    name: str
    show_gridlines: bool = True
    cells: dict[tuple[int, int], CellLayout] = field(default_factory=dict)
    covered_cells: set[tuple[int, int]] = field(default_factory=set)
    column_widths: dict[int, float] = field(default_factory=dict)
    row_heights: dict[int, float] = field(default_factory=dict)
    background_color: str | None = None
    max_row: int = 0
    max_col: int = 0


@dataclass(frozen=True)
class WorkbookLayout:
    """A workbook prepared for PDF or image rendering."""

    sheets: tuple[SheetLayout, ...]


class _LayoutEngine(Engine):
    def __init__(self) -> None:
        super().__init__()
        self.sheets: list[SheetLayout] = []
        self._current_sheet: SheetLayout | None = None

    def create_sheet(self, name: str, show_gridlines: bool = True) -> None:
        sheet = SheetLayout(name=name, show_gridlines=show_gridlines)
        self.sheets.append(sheet)
        self._current_sheet = sheet

    def write_cell(
        self,
        row: int,
        col: int,
        value: object,
        style: EffectiveStyle,
        border_fallback_color: str,
    ) -> None:
        self._add_cell(row, col, 1, 1, value, style, border_fallback_color)

    def write_merged_cell(
        self,
        row: int,
        col: int,
        rowspan: int,
        colspan: int,
        value: object,
        style: EffectiveStyle,
        border_fallback_color: str,
    ) -> None:
        self._add_cell(
            row,
            col,
            rowspan,
            colspan,
            value,
            style,
            border_fallback_color,
        )

    def set_column_width(self, col: int, width: float) -> None:
        sheet = self._require_sheet()
        sheet.column_widths[col] = width
        sheet.max_col = max(sheet.max_col, col)

    def set_row_height(self, row: int, height: float) -> None:
        sheet = self._require_sheet()
        sheet.row_heights[row] = height
        sheet.max_row = max(sheet.max_row, row)

    def fill_background(self, color: str, max_row: int, max_col: int) -> None:
        # Exporters paint the finite report canvas rather than materializing
        # Excel's large background-fill range.
        self._require_sheet().background_color = normalize_hex(color)

    def copy_sheet(
        self,
        source: SaveTarget | bytes | BinaryIO,
        sheet_name: str,
        dest_name: str,
        show_gridlines: bool | None = None,
    ) -> None:
        workbook = _load_source_workbook(source)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in source workbook")
        source_sheet = workbook[sheet_name]
        source_gridlines = source_sheet.sheet_view.showGridLines
        self.create_sheet(
            dest_name,
            show_gridlines=(True if source_gridlines is None else source_gridlines)
            if show_gridlines is None
            else show_gridlines,
        )

        merges: dict[tuple[int, int], tuple[int, int]] = {}
        for merged_range in source_sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            rowspan = max_row - min_row + 1
            colspan = max_col - min_col + 1
            merges[(min_row, min_col)] = (rowspan, colspan)

        for row in source_sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value is None and not getattr(cell, "has_style", False):
                    if (cell.row, cell.column) not in merges:
                        continue
                rowspan, colspan = merges.get((cell.row, cell.column), (1, 1))
                self._add_cell(
                    cell.row,
                    cell.column,
                    rowspan,
                    colspan,
                    cell.value,
                    _style_from_openpyxl_cell(cell),
                    _DEFAULT_BORDER_COLOR,
                )

        for column_letter, dimension in source_sheet.column_dimensions.items():
            if dimension.width:
                self.set_column_width(
                    _column_letter_to_index(column_letter), float(dimension.width)
                )
        for row_index, dimension in source_sheet.row_dimensions.items():
            if dimension.height:
                self.set_row_height(row_index, float(dimension.height))

    def save(self, target: SaveTarget | None = None) -> bytes | None:
        raise RuntimeError("Layout engines cannot be saved directly")

    def _add_cell(
        self,
        row: int,
        col: int,
        rowspan: int,
        colspan: int,
        value: object,
        style: EffectiveStyle,
        border_fallback_color: str,
    ) -> None:
        sheet = self._require_sheet()
        sheet.cells[(row, col)] = CellLayout(
            row=row,
            col=col,
            value=normalize_cell_value(value),
            style=style,
            border_fallback_color=normalize_hex(border_fallback_color),
            rowspan=rowspan,
            colspan=colspan,
        )
        for row_index in range(row, row + rowspan):
            for column_index in range(col, col + colspan):
                if (row_index, column_index) != (row, col):
                    sheet.covered_cells.add((row_index, column_index))
        sheet.max_row = max(sheet.max_row, row + rowspan - 1)
        sheet.max_col = max(sheet.max_col, col + colspan - 1)

    def _require_sheet(self) -> SheetLayout:
        if self._current_sheet is None:
            raise RuntimeError("No sheet has been created")
        return self._current_sheet


def build_workbook_layout(node: WorkbookNode) -> WorkbookLayout:
    """Resolve a workbook node into the shared export layout."""
    engine = _LayoutEngine()
    for sheet in node.sheets:
        render_sheet(engine, sheet)
    return WorkbookLayout(tuple(engine.sheets))


def _load_source_workbook(source: SaveTarget | bytes | BinaryIO):
    if isinstance(source, (str, Path)):
        return load_workbook(source, data_only=False)
    if isinstance(source, bytes):
        stream: BinaryIO = BytesIO(source)
    else:
        stream = source
        try:
            stream.seek(0)
        except (AttributeError, OSError):
            stream = BytesIO(stream.read())
    return load_workbook(stream, data_only=False)


def _style_from_openpyxl_cell(cell: object) -> EffectiveStyle:
    font = cell.font  # type: ignore[attr-defined]
    fill = cell.fill  # type: ignore[attr-defined]
    alignment = cell.alignment  # type: ignore[attr-defined]
    border = cell.border  # type: ignore[attr-defined]
    text_color = _openpyxl_color_to_hex(font.color) or _DEFAULT_TEXT_COLOR
    fill_color = None
    if getattr(fill, "fill_type", None) == "solid":
        fill_color = _openpyxl_color_to_hex(fill.fgColor)

    border_style: BorderStyleName | None = None
    border_color = None
    present_sides: set[str] = set()
    for side_name in ("top", "bottom", "left", "right"):
        side = getattr(border, side_name, None)
        if not side or not side.style or side.style == "none":
            continue
        present_sides.add(side_name)
        if border_style is None:
            border_style = side.style
        if border_color is None:
            border_color = _openpyxl_color_to_hex(side.color)

    return EffectiveStyle(
        font_name=font.name or _DEFAULT_FONT_NAME,
        font_size=float(font.size or _DEFAULT_FONT_SIZE),
        bold=bool(font.bold),
        italic=bool(font.italic),
        text_color=text_color,
        fill_color=fill_color,
        horizontal_align=alignment.horizontal,
        vertical_align=alignment.vertical,
        indent=int(alignment.indent) if alignment.indent else None,
        wrap_text=bool(alignment.wrap_text),
        shrink_to_fit=bool(alignment.shrink_to_fit),
        auto_width=True,
        row_height=None,
        row_width=None,
        number_format=cell.number_format if cell.number_format != "General" else None,  # type: ignore[attr-defined]
        border=border_style,
        border_color=border_color,
        border_top="top" in present_sides,
        border_bottom="bottom" in present_sides,
        border_left="left" in present_sides,
        border_right="right" in present_sides,
    )


def _openpyxl_color_to_hex(color: object | None) -> str | None:
    rgb = getattr(color, "rgb", None)
    if isinstance(rgb, str) and len(rgb) >= 6:
        return normalize_hex("#" + rgb[-6:])
    return None


def _column_letter_to_index(letter: str) -> int:
    result = 0
    for character in letter.upper():
        result = result * 26 + ord(character) - ord("A") + 1
    return result
