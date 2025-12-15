"""OpenPyXL rendering engine implementation."""

from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any, BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..styles import to_argb
from .base import EffectiveStyle, Engine, SaveTarget

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

__all__ = ["OpenpyxlEngine"]


@dataclass
class _TemplateState:
    source_identity: tuple[str, str] | None
    keep_sheets: set[str] = field(default_factory=set)

    def keep(self, name: str) -> None:
        self.keep_sheets.add(name)


class OpenpyxlEngine(Engine):
    """Rendering engine using openpyxl."""

    def __init__(self) -> None:
        super().__init__()
        self._workbook = Workbook()
        # Remove default sheet created by openpyxl
        default_sheet = self._workbook.active
        if default_sheet is not None:
            self._workbook.remove(default_sheet)
        self._current_sheet: Worksheet | None = None
        self._template: _TemplateState | None = None
        # Cache style objects to avoid duplicates
        self._style_cache: dict[tuple[Any, ...], dict[str, Any]] = {}
        # Cache color conversions
        self._color_cache: dict[str, str] = {}
        # Cache column letters
        self._column_letter_cache: dict[int, str] = {}

    @classmethod
    def from_workbook(cls, workbook: Workbook) -> "OpenpyxlEngine":
        """Create an engine instance wrapping an existing openpyxl Workbook.

        This allows reusing copy_sheet() and other engine methods against
        a workbook that was loaded or created externally (e.g., from xlsxwriter bytes).
        """
        engine = cls.__new__(cls)
        engine._workbook = workbook
        engine._current_sheet = None
        engine._template = None
        engine._style_cache = {}
        engine._color_cache = {}
        engine._column_letter_cache = {}
        return engine

    def create_sheet(self, name: str) -> None:
        self._current_sheet = self._workbook.create_sheet(title=name)
        if self._template is not None:
            self._template.keep(name)

    def write_cell(
        self,
        row: int,
        col: int,
        value: object,
        style: EffectiveStyle,
        border_fallback_color: str,
    ) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        cell = self._current_sheet.cell(row=row, column=col)
        cell.value = value  # type: ignore[assignment]
        self._apply_style(cell, style, border_fallback_color)

    def _get_cached_color(self, color: str) -> str:
        """Get cached ARGB color or convert and cache it."""
        if color not in self._color_cache:
            self._color_cache[color] = to_argb(color)
        return self._color_cache[color]

    def _get_cached_styles(
        self, effective: EffectiveStyle, border_fallback_color: str
    ) -> dict[str, Any]:
        """Get or create cached style objects for the given style."""
        # Create a hashable key from style properties
        cache_key = (
            effective.font_name,
            effective.font_size,
            effective.bold,
            effective.italic,
            effective.text_color,
            effective.fill_color,
            effective.horizontal_align,
            effective.vertical_align,
            effective.indent,
            effective.wrap_text,
            effective.shrink_to_fit,
            effective.number_format,
            effective.border,
            effective.border_color or border_fallback_color
            if effective.border
            else None,
            effective.border_top,
            effective.border_bottom,
            effective.border_left,
            effective.border_right,
        )

        if cache_key in self._style_cache:
            return self._style_cache[cache_key]

        # Create style objects
        text_color_argb = self._get_cached_color(effective.text_color)
        font = Font(
            name=effective.font_name,
            size=effective.font_size,
            bold=effective.bold,
            italic=effective.italic,
            color=text_color_argb,
        )

        fill: PatternFill | None = None
        if effective.fill_color:
            fill_color_argb = self._get_cached_color(effective.fill_color)
            fill = PatternFill(
                fill_type="solid",
                start_color=fill_color_argb,
                end_color=fill_color_argb,
            )

        alignment: Alignment | None = None
        align_kwargs: dict[str, Any] = {}
        if effective.horizontal_align:
            align_kwargs["horizontal"] = effective.horizontal_align
        if effective.vertical_align:
            align_kwargs["vertical"] = effective.vertical_align
        if effective.indent is not None:
            align_kwargs["indent"] = effective.indent
        if effective.wrap_text:
            align_kwargs["wrap_text"] = True
        if effective.shrink_to_fit:
            align_kwargs["shrink_to_fit"] = True
        if align_kwargs:
            align_kwargs.setdefault("vertical", "bottom")
            alignment = Alignment(**align_kwargs)  # type: ignore[arg-type]
        elif effective.wrap_text or effective.shrink_to_fit:
            alignment = Alignment(
                wrap_text=True if effective.wrap_text else None,
                shrink_to_fit=True if effective.shrink_to_fit else None,
            )

        border: Border | None = None
        if effective.border:
            border_color = effective.border_color or border_fallback_color
            border_color_argb = self._get_cached_color(border_color)

            def build_side(enabled: bool) -> Side | None:
                if not enabled:
                    return None
                return Side(style=effective.border, color=border_color_argb)

            explicit = (
                effective.border_top
                or effective.border_bottom
                or effective.border_left
                or effective.border_right
            )
            if explicit:
                border = Border(
                    left=build_side(effective.border_left),
                    right=build_side(effective.border_right),
                    top=build_side(effective.border_top),
                    bottom=build_side(effective.border_bottom),
                )
            else:
                side = build_side(True)
                border = Border(left=side, right=side, top=side, bottom=side)

        styles = {
            "font": font,
            "fill": fill,
            "alignment": alignment,
            "border": border,
            "number_format": effective.number_format,
        }

        self._style_cache[cache_key] = styles
        return styles

    def _apply_style(
        self, cell: object, effective: EffectiveStyle, border_fallback_color: str
    ) -> None:
        """Apply style to an openpyxl cell."""
        styles = self._get_cached_styles(effective, border_fallback_color)

        cell.font = styles["font"]  # type: ignore[attr-defined]

        if styles["fill"]:
            cell.fill = styles["fill"]  # type: ignore[attr-defined]

        if styles["alignment"]:
            cell.alignment = styles["alignment"]  # type: ignore[attr-defined]
        elif cell.alignment is None and (  # type: ignore[attr-defined]
            effective.wrap_text or effective.shrink_to_fit
        ):
            # Handle edge case where alignment wasn't cached but wrap/shrink is needed
            cell.alignment = Alignment(  # type: ignore[attr-defined]
                wrap_text=True if effective.wrap_text else None,
                shrink_to_fit=True if effective.shrink_to_fit else None,
            )

        if styles["number_format"]:
            cell.number_format = styles["number_format"]  # type: ignore[attr-defined]

        if styles["border"]:
            cell.border = styles["border"]  # type: ignore[attr-defined]

    def set_column_width(self, col: int, width: float) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        if col not in self._column_letter_cache:
            self._column_letter_cache[col] = get_column_letter(col)
        letter = self._column_letter_cache[col]
        self._current_sheet.column_dimensions[letter].width = max(width, 8.0)

    def set_row_height(self, row: int, height: float) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        self._current_sheet.row_dimensions[row].height = height

    def fill_background(
        self,
        color: str,
        max_row: int,
        max_col: int,
    ) -> None:
        if self._current_sheet is None:
            raise RuntimeError("No sheet created. Call create_sheet() first.")

        fill_color = self._get_cached_color(color)
        sheet_fill = PatternFill(
            fill_type="solid", start_color=fill_color, end_color=fill_color
        )
        # Reuse the same PatternFill object for all cells (openpyxl supports this)
        for row in self._current_sheet.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=max_col
        ):
            for cell in row:
                cell.fill = sheet_fill

    def _deep_copy_workbook(self, workbook: Workbook) -> Workbook:
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return load_workbook(
            buffer,
            data_only=False,
        )

    def _clone_sheet_contents(self, source_ws: Worksheet, target_ws: Worksheet) -> None:
        """Copy values and basic layout from a source sheet into a target sheet.

        This is intentionally minimal to avoid copying workbook-scoped internals
        (e.g. style indices) that can corrupt files when saving after a cross-
        workbook copy.
        """
        # Values (including formulas); no styles.
        for row in source_ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.row is None or cell.column is None:
                    continue
                target_ws.cell(row=cell.row, column=cell.column, value=cell.value)

        # Merge ranges
        for merged_range in source_ws.merged_cells.ranges:
            target_ws.merge_cells(str(merged_range))

        # Column widths
        for column, dimension in source_ws.column_dimensions.items():
            if getattr(dimension, "width", None):
                target_ws.column_dimensions[column].width = dimension.width

        # Row heights
        for row, dimension in source_ws.row_dimensions.items():
            if getattr(dimension, "height", None):
                target_ws.row_dimensions[row].height = dimension.height

        # Freeze panes
        target_ws.freeze_panes = source_ws.freeze_panes

        # Auto filter
        try:
            target_ws.auto_filter.ref = source_ws.auto_filter.ref
        except Exception:
            pass

        # Zoom scale
        try:
            target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
        except Exception:
            pass

    def _source_identity(
        self, source: SaveTarget | bytes | BinaryIO
    ) -> tuple[str, str] | None:
        if isinstance(source, (str, Path)):
            return ("path", str(Path(source).resolve()))
        if isinstance(source, bytes):
            return ("bytes", hashlib.sha256(source).hexdigest())
        return None

    def _load_source_workbook(
        self, source: SaveTarget | bytes | BinaryIO
    ) -> tuple[Workbook, tuple[str, str] | None]:
        identity = self._source_identity(source)

        if isinstance(source, (str, Path)):
            workbook = load_workbook(
                filename=source,
                data_only=False,
            )
            return workbook, identity

        buffer: BinaryIO
        if isinstance(source, bytes):
            buffer = BytesIO(source)
        else:
            buffer = source
            if hasattr(buffer, "seek"):
                try:
                    buffer.seek(0)
                except Exception:
                    pass

        workbook = load_workbook(
            buffer,
            data_only=False,
        )
        return workbook, identity

    def _ensure_sheet_name_available(self, name: str) -> None:
        if name in self._workbook.sheetnames:
            raise ValueError(f"Sheet '{name}' already exists in destination workbook")

    def _maybe_start_from_template(
        self, source_wb: Workbook, source_identity: tuple[str, str] | None
    ) -> None:
        if self._template is not None:
            return
        if self._workbook.sheetnames:
            return

        # Prefer a workbook-level deep copy when we haven't rendered anything yet.
        # This keeps themes, shared styles, and other workbook-level structures intact.
        self._template = _TemplateState(source_identity=source_identity)
        self._workbook = self._deep_copy_workbook(source_wb)
        self._current_sheet = None

    def _resolve_copy_source_sheet(
        self,
        source_wb: Workbook,
        source_identity: tuple[str, str] | None,
        sheet_name: str,
    ) -> Worksheet:
        if (
            self._template is not None
            and self._template.source_identity is not None
            and self._template.source_identity == source_identity
            and sheet_name in self._workbook.sheetnames
        ):
            return self._workbook[sheet_name]
        return source_wb[sheet_name]

    def copy_sheet(
        self, source: SaveTarget | bytes | BinaryIO, sheet_name: str, dest_name: str
    ) -> None:
        """Copy a sheet from an external workbook without translation."""

        source_wb, source_identity = self._load_source_workbook(source)

        if sheet_name not in source_wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in source workbook")

        self._ensure_sheet_name_available(dest_name)

        self._maybe_start_from_template(source_wb, source_identity)

        source_ws = self._resolve_copy_source_sheet(
            source_wb, source_identity, sheet_name
        )
        if source_ws.parent is self._workbook:
            target_ws = self._workbook.copy_worksheet(source_ws)
            target_ws.title = dest_name
        else:
            target_ws = self._workbook.create_sheet(title=dest_name)
            self._clone_sheet_contents(source_ws, target_ws)

        self._current_sheet = target_ws
        if self._template is not None:
            self._template.keep(dest_name)

    def save(self, target: SaveTarget | None = None) -> bytes | None:
        if self._template is not None and self._template.keep_sheets:
            for ws in list(self._workbook.worksheets):
                if ws.title not in self._template.keep_sheets:
                    self._workbook.remove(ws)

        if target is None:
            buffer = BytesIO()
            self._workbook.save(buffer)
            return buffer.getvalue()

        if isinstance(target, (str, Path)):
            self._workbook.save(str(target))
        else:
            self._workbook.save(target)
        return None
